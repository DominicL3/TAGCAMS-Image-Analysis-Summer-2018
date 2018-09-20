'''Combines functions from both the 6month data and 18month data. Be VERY CAREFUL when assigning variables
and be sure that you aren't confusing variables from tagcams.py with those from osiris18month.py'''
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import pandas as pd
from astropy.io import fits
from astropy.modeling import models, fitting
from skimage import transform
from sklearn.linear_model import LinearRegression
from matplotlib.backends.backend_pdf import PdfPages
import os
import glob

plt.ion()
plt.style.use('bmh')
mpl.rcParams['axes.titlesize'] = 12
mpl.rcParams['axes.labelsize'] = 12
mpl.rcParams['xtick.labelsize'] = 9
mpl.rcParams['ytick.labelsize'] = 9
mpl.rcParams['font.family'] = 'Myriad Pro'
plt.rcParams['mathtext.fontset'] = 'custom'
plt.rcParams['mathtext.rm'] = 'Myriad Pro'

# Allows you to use functions from tagcams.py and osiris18month.py just by calling tagcams.function! I didn't know this before!
import tagcams
import osiris18month
from cycler import cycler
default_cycler = cycler('color', ['tab:red', 'tab:blue', 'orange', 'tab:green', 'tab:purple'])
mpl.rcParams['axes.prop_cycle'] = default_cycler

def MTFs_6month(region='center', DOY=103, camera='NFT', gain=8):
    [half_inds_DOYgaincamera, regions_DOYgaincamera] = tagcams.visualize_PSFs(DOY=DOY, camera=camera, gain=gain, plot=False)[2]

    half_inds_DOYgaincamera, regions_DOYgaincamera = np.array(half_inds_DOYgaincamera), np.array(regions_DOYgaincamera)

    # retrieves the indices where only if the regions match the input region
    filtered_indices_6month = half_inds_DOYgaincamera[np.where(regions_DOYgaincamera == region)[0]]

    hybrids_6month = [tagcams.hybrid_PSF(index) for index in filtered_indices_6month]
    horizontal_MTFs, vertical_MTFs = [], []

    for hybrid in hybrids_6month:
        MTF = tagcams.calculate_MTF(hybrid)
        horizontal_MTFs.append(MTF[0, :5])
        vertical_MTFs.append(MTF[:5, 0])
        assert len(horizontal_MTFs) == len(vertical_MTFs)

    return horizontal_MTFs, vertical_MTFs

def MTFs_18month(region='center', camera='NFT', gain=8):
    [half_inds_gaincamera, regions_gaincamera] = osiris18month.visualize_PSFs(camera=camera, gain=gain, plot=False)[2]

    half_inds_gaincamera, regions_gaincamera = np.array(half_inds_gaincamera), np.array(regions_gaincamera)

    # retrieves the indices where only if the regions match the input region
    filtered_indices_18month = half_inds_gaincamera[np.where(regions_gaincamera == region)[0]]

    hybrids_18month = [osiris18month.hybrid_PSF(index) for index in filtered_indices_6month]
    horizontal_MTFs, vertical_MTFs = [], []

    for hybrid in hybrids_18month:
        MTF = osiris18month.calculate_MTF(hybrid)
        horizontal_MTFs.append(MTF[0, :5])
        vertical_MTFs.append(MTF[:5, 0])
        assert len(horizontal_MTFs) == len(vertical_MTFs)

    return horizontal_MTFs, vertical_MTFs

def faster_MTFs_6month(region='center', camera='NCM', gain=8):
    # filters the 0.5s and 5s images by gain and camera, but not by DOY; gets image names
    filtered_half_images = [img for img in tagcams.half_second_images if fits.getheader(img)['TCGAIN']==gain and fits.getheader(img)['INSTRUME']==camera and img not in np.array(tagcams.half_second_images)[tagcams.washed_out_images]]
    filtered_five_images = [img for img in tagcams.five_second_images if fits.getheader(img)['TCGAIN']==gain and fits.getheader(img)['INSTRUME']==camera and img not in np.array(tagcams.five_second_images)[tagcams.washed_out_images]]

    # get indices of all the pics in the filtered images
    half_indexes = [tagcams.half_second_images.index(img) for img in filtered_half_images]
    five_indexes = [tagcams.five_second_images.index(img) for img in filtered_five_images]

    # determines the region that each of these images are in
    regions_halfsec = [tagcams.assign_region(tagcams.brights_halfsec[index]) for index in half_indexes]
    regions_fivesec = [tagcams.assign_region(tagcams.brights_5s[index]) for index in five_indexes]
    assert regions_halfsec==regions_fivesec, 'Regions are not the same!'

    half_indexes, regions_halfsec = np.array(half_indexes), np.array(regions_halfsec)
    assert len(half_indexes) == len(regions_halfsec), 'Index lengths and regions list are not the same'

    # retrieves the indices where only if the regions match the input region
    filtered_indices_by_region = half_indexes[regions_halfsec == region]

    hybrids_6month = [tagcams.hybrid_PSF(index) for index in filtered_indices_by_region]
    horizontal_MTFs, vertical_MTFs, DOY_list = [], [], []
    for index in filtered_indices_by_region:
        if tagcams.half_second_images[index] in tagcams.globs[:58]:
            DOY_list.append('100')
        elif tagcams.half_second_images[index] in tagcams.globs[58:]:
            DOY_list.append('103')
        else:
            return "Invalid index. Can't match to DOY"
    assert len(DOY_list) == len(filtered_indices_by_region)

    for hybrid in hybrids_6month:
        MTF = tagcams.calculate_MTF(hybrid)
        horizontal_MTFs.append(MTF[0, :5])
        vertical_MTFs.append(MTF[:5, 0])
        assert len(horizontal_MTFs) == len(vertical_MTFs)

    return horizontal_MTFs, vertical_MTFs, DOY_list

def faster_MTFs_18month(region='center', camera='NCM', gain=8):
    # filters the 0.5s and 5s images by gain and camera, but not by DOY; gets image names
    filtered_half_images = [img for img in osiris18month.half_second_images if fits.getheader(img)['TCGAIN']==gain and fits.getheader(img)['INSTRUME']==camera and img not in np.array(osiris18month.half_second_images)[osiris18month.washed_out_images]]
    filtered_five_images = [img for img in osiris18month.five_second_images if fits.getheader(img)['TCGAIN']==gain and fits.getheader(img)['INSTRUME']==camera and img not in np.array(osiris18month.five_second_images)[osiris18month.washed_out_images]]

    # get indices of all the pics in the filtered images
    half_indexes = [osiris18month.half_second_images.index(img) for img in filtered_half_images]
    five_indexes = [osiris18month.five_second_images.index(img) for img in filtered_five_images]

    # determines the region that each of these images are in
    regions_halfsec = [osiris18month.assign_region(osiris18month.brights_halfsec[index]) for index in half_indexes]
    regions_fivesec = [osiris18month.assign_region(osiris18month.brights_5s[index]) for index in five_indexes]
    assert regions_halfsec==regions_fivesec, 'Regions are not the same!'

    half_indexes, regions_halfsec = np.array(half_indexes), np.array(regions_halfsec)
    assert len(half_indexes) == len(regions_halfsec), 'Index lengths and regions list are not the same'

    # retrieves the indices where only if the regions match the input region
    filtered_indices_by_region = half_indexes[regions_halfsec == region]

    hybrids_18month = [osiris18month.hybrid_PSF(index) for index in filtered_indices_by_region]
    horizontal_MTFs, vertical_MTFs = [], []

    for hybrid in hybrids_18month:
        MTF = osiris18month.calculate_MTF(hybrid)
        horizontal_MTFs.append(MTF[0, :5])
        vertical_MTFs.append(MTF[:5, 0])
        assert len(horizontal_MTFs) == len(vertical_MTFs)

    return horizontal_MTFs, vertical_MTFs

def sameplot_MTF(MTF_direction, region='center', camera='NCM', gain=8):
    '''Puts the horizontal and vertical MTFs from 6months and 18months on the same plot'''
    # loads in the MTFs from both regions
    # honestly would be much quicker to do all regions at once but who cares
    horizontal_MTF6months, vertical_MTF6months, DOY_list = faster_MTFs_6month(region=region, camera=camera, gain=gain)
    horizontal_MTF18months, vertical_MTF18months= faster_MTFs_18month(region=region, camera=camera, gain=gain)

    # narrows down the MTF to be used to either the horizontal or vertical version
    if MTF_direction == 'horizontal':
        MTF6months, MTF18months = horizontal_MTF6months, horizontal_MTF18months
    elif MTF_direction == 'vertical':
        MTF6months, MTF18months = vertical_MTF6months, vertical_MTF18months
    else:
        return 'Not a valid version of the MTF'

    fig, ax = plt.subplots(1,1, figsize=(10,8))
    spatial_frequencies = np.arange(5) / (10 * 2.2e-3) # n/(10 pixels * 2.2e-3mm) to get units of 1/mm
    for DOY, MTF in zip(DOY_list, MTF6months):
        ax.plot(spatial_frequencies, MTF, linestyle='--', marker='o', label='L+6Months (DOY{})'.format(DOY))
    for MTF in MTF18months:
        ax.plot(spatial_frequencies, MTF, marker='o', label='L+18Months')

    ax.set(xlabel='Spatial Frequency ' + r'$\mathrm{[mm^{-1}]}$',ylabel='{} MTF'.format(MTF_direction.title()))
    ax.set_title('Region: {}'.format(region.title()), loc='center')
    ax.set_title('Gain: {}'.format(gain), loc='right')
    ax.set_title('Camera: {}'.format(camera), loc='left')
    #ax.xaxis.set_ticklabels([]), ax.yaxis.set_ticklabels([])
    ax.legend()
    fig.tight_layout()
    return fig, ax

def produce_MTFpngs(camera='NCM', gain=8):
    plt.ioff()
    for MTF_direction in ['horizontal', 'vertical']:
        for region in tagcams.region_colors.keys():
            f = sameplot_MTF(MTF_direction, region=region, camera=camera, gain=gain)[0]
            f.savefig('MTF comparisons by region pngs/NFT combined DOY/{0}, region={1}, camera={2}, gain{3}.png'.format(MTF_direction, region, camera, gain), dpi=500)
            plt.close(f)

def region_rectangles(save=True):
    sample_image = fits.getdata(osiris18month.five_second_images[31])
    w, h = mpl.figure.figaspect(sample_image)
    for region in osiris18month.flipped_rectangle_centers.keys():
        f,ax = plt.subplots(1,1,figsize=(w,h))
        ax.imshow(sample_image, vmin=0, vmax=30, cmap='gray')
        ax.axis('off')
        #ax.grid(False)
        brightest_row, brightest_column = osiris18month.flipped_rectangle_centers[region]
        rect_color = osiris18month.region_colors[region]
        rectangle = patches.Rectangle([brightest_column-65,brightest_row-50], 130, 100, edgecolor=rect_color,linewidth=0.7,facecolor='none')
        _= ax.add_patch(rectangle)
        #ax.set(xticks=[], yticks=[])
        f.tight_layout()
        f.subplots_adjust(left=0,right=1,bottom=0,top=1)
        if save:
            f.savefig('MTF comparisons by region pngs/single rectangle over regions/{}.png'.format(region),pad_inches=0, dpi=300)


def plot_MTFs_over_FOV(MTF_direction, camera, gain, image_divisor, vmin_vmax=[0,30]):
    '''Creates MTFs across the field of view.'''
    # plots the sample image
    fig, ax = plt.subplots(1,1,figsize=(12,9))
    sample_image_file = osiris18month.five_second_images[31]
    ax.imshow(fits.getdata(sample_image_file),vmin=vmin_vmax[0],vmax=vmin_vmax[1],cmap='gray')

    folder = 'MTF comparisons by region pngs/{}, gain{}'.format(camera, gain)
    pngs = [file for file in os.listdir(folder) if MTF_direction in file]
    assert 2400 % image_divisor == 0
    assert 3200 % image_divisor == 0
    assert len(pngs) == 13, 'Missing plots'

    box_half = np.array([2400, 3200]) / image_divisor
    for region in osiris18month.flipped_rectangle_centers.keys():
        matching_region_files = [s for s in pngs if region in s]
        check_regions_list = ['corner', 'top', 'bottom']
        for r in check_regions_list:
            if r not in region:
                matching_region_files = [m for m in matching_region_files if r not in m]
        if region == 'top' or region == 'bottom':
            check_regions_list2 = ['left', 'right']
            for r in check_regions_list2:
                matching_region_files = [m for m in matching_region_files if r not in m]
        print (region)
        assert len(matching_region_files) == 1, 'More string matches!'

        matching_file = matching_region_files[0]
        PSF_location = osiris18month.flipped_rectangle_centers[region]

        # plot the image on the FOV
        image_data = plt.imread('{}/{}'.format(folder, matching_file))
        # sets location of the image using extent (left, right, bottom, top)
        extent = [PSF_location[1]-box_half[1], PSF_location[1]+box_half[1], PSF_location[0]+box_half[0], PSF_location[0]-box_half[0]]
        ax.imshow(image_data, extent=extent)
    # needs to be set as the last line to display the full image, otherwise will only display the last imshow instance
    ax.set(xlim=[0,2752],ylim=(2004,0), xticks=[],yticks=[])
    ax.grid(False)
    fig.tight_layout()

def NFT_coordinates(old_coordinates):
    '''Accepts an array and converts it to the NFT coordinate system.
    ASSUMES that old_coordinates are already given in Python (x,y) format, not (row, column)'''
    X_DL, Y_DL = old_coordinates
    return (2731 - X_DL, Y_DL - 54)

def excel_bookkeeping_actualdata(output_name, write=True):
    '''Creates an Excel spreadsheet that summarizes some data from the L+6Months and L+18Month images.
    Columns include filename, camera, bitmode, exposure, and brightest pixel.'''
    dataframes = []
    for dataset in [tagcams, osiris18month]:
        image_names = [file[34:] for file in dataset.globs]
        cameras = [fits.getheader(img)['INSTRUME'] for img in dataset.globs]
        gains = [fits.getheader(img)['TCGAIN'] for img in dataset.globs]
        bitmodes = [fits.getheader(img)['TCBPP'] for img in dataset.globs]
        exposures = [fits.getheader(img)['EXPTIME'] for img in dataset.globs]

        bright_pixels_reversed_halfsec = [(int(column_row[1]), int(column_row[0])) for column_row in dataset.brights_halfsec]
        bright_pixels_reversed_fivesec = [(int(column_row[1]), int(column_row[0])) for column_row in dataset.brights_5s]
        region_from_data = [dataset.assign_region(pixels) for pixels in dataset.brights_5s]

        # changes to coordinate system used by other OSIRIS-REx folks
        NFT_halfsec_pixels = [NFT_coordinates(old_pixels) for old_pixels in bright_pixels_reversed_halfsec]
        NFT_fivesec_pixels = [NFT_coordinates(old_pixels) for old_pixels in bright_pixels_reversed_fivesec]

        # will be used to iterate through the bright pixels in the 0.5s and 5s images
        target_coordinates, targets_python, regions = [], [], []
        i, j = 0, 0

        for exp in exposures:
            if np.round(exp,2) == 0.5:
                target_coordinates.append(NFT_halfsec_pixels[i])
                targets_python.append(bright_pixels_reversed_halfsec[i])
                regions.append(region_from_data[i])
                i += 1
            elif np.round(exp,2) == 5:
                target_coordinates.append(NFT_fivesec_pixels[j])
                targets_python.append(bright_pixels_reversed_fivesec[j])
                regions.append(region_from_data[j])
                j += 1
            else:
                target_coordinates.append(np.nan)
                targets_python.append(np.nan)
                regions.append(np.nan)
        assert len(target_coordinates) == len(dataset.globs)
        assert len(regions) == len(dataset.globs)

        # Create a pandas Dataframe for Excel spreadsheet
        data = np.array([image_names, cameras, exposures, gains, bitmodes, target_coordinates, targets_python, regions], dtype=object).T
        df = pd.DataFrame(data, columns=['Filename', 'Camera', 'Exposure', 'Gain', 'Bitmode', 'Brightest Pixel of Target (NFT coordinates)', 'Brightest Pixel of Target (Python coordinates)', 'FOV Region (Python coordinates)'])
        dataframes.append(df)

    df_6month, df_18month = dataframes
    if write:
        with pd.ExcelWriter(output_name, engine='xlsxwriter') as writer:
            df_6month.to_excel(writer, sheet_name ='L+6Month')
            df_18month.to_excel(writer, sheet_name = 'L+18Month')
            writer.save()
    return df_6month, df_18month

#excel_bookkeeping_actualdata('L+6 and L+18 with NFT coordinates.xlsx')

def export_hybrids(excel_name, clobber=True):
    '''Write all the hybrid PSFs from the L+6Month and L+18Month images to FITS files.
    Check out if the arrangement of arrays looks okay.'''
    img_analysis_scripts = [tagcams, osiris18month]
    from openpyxl import load_workbook
    existing_workbook = load_workbook('L+6 and L+18 images summary.xlsx')

    for script, dataframe in zip(img_analysis_scripts, excel_bookkeeping_actualdata('fakename', write=False)):
        # pulls DataFrame from L+6Month and L+18, deletes nan rows, and sorts by exposure time and removes indices
        half_second_df = dataframe.dropna().loc[np.where(np.round(dataframe['Exposure'].values.astype(float), 2)==0.5)].reset_index()
        fivesecond_df = dataframe.dropna().loc[np.where(np.round(dataframe['Exposure'].values.astype(float), 2)==5)].reset_index()

        # rename the columns and drop a few redundant ones in the 0.5s DataFrame
        half_second_df = half_second_df.rename(columns={'index':'File Index (0.5s)','Filename':'Filename (0.5s)', \
        'Brightest Pixel of Target (NFT coordinates)':'Brightest Pixel of Target (NFT) (0.5s)',\
        'Brightest Pixel of Target (Python coordinates)': 'Brightest Pixel of Target (Python) (0.5s)'}).drop(['Camera', 'Exposure', 'Gain', 'Bitmode', 'FOV Region (Python coordinates)'], axis=1)
        fivesecond_df = fivesecond_df.rename(columns={'index':'File Index (5s)','Filename':'Filename (5s)', \
        'Brightest Pixel of Target (NFT coordinates)':'Brightest Pixel of Target (NFT) (5s)', \
        'Brightest Pixel of Target (Python coordinates)': 'Brightest Pixel of Target (Python) (5s)'}).drop(['Exposure'], axis=1)

        combined_images = pd.concat([half_second_df,fivesecond_df], axis=1)
        column_order = [0,1,4,5,6,7,8,2,9,3,10,11]
        combined_images = combined_images[[combined_images.columns[i] for i in column_order]]

        # puts this nice looking DataFrame into an excel file
        with pd.ExcelWriter(excel_name, engine='openpyxl') as writer:
            writer.book = existing_workbook
            if script == tagcams:
                sheet = 'L+6 Hybrids'
            elif script == osiris18month:
                sheet = 'L+18 Hybrids'
            combined_images.to_excel(writer, sheet_name=sheet)

        # retrieves each row in the DataFrame and computes the hybrid PSF
        for row in np.arange(len(combined_images)):
            if row in script.washed_out_images: # filters out the washed_out_images
                continue
            piece_of_data = combined_images.iloc[row]
            camera, gain, bitmode, region = piece_of_data[['Camera','Gain', 'Bitmode', 'FOV Region (Python coordinates)']]
            hybrid = script.hybrid_PSF(row)

            # creates fits file with the data and header information
            fits_file = fits.PrimaryHDU()
            fits_file.data = hybrid
            for header_title in piece_of_data.index:
                if type(piece_of_data[header_title]) == tuple: # since FITS headers can't read tuples, put tuples in comments section
                    fits_file.header[header_title] = piece_of_data[header_title]
                    fits_file.header.comments[header_title] = piece_of_data[header_title]
                else:
                    fits_file.header[header_title] = piece_of_data[header_title]

            subfolder = '/gain {0}/bitmode {1}/{5}'.format(gain, bitmode, piece_of_data['File Index (0.5s)'], piece_of_data['File Index (5s)'], camera, region)
            if script == tagcams:
                full_folder_path = 'hybrids data/L+6Month {}'.format(camera) + subfolder
            elif script == osiris18month:
                full_folder_path = 'hybrids data/L+18Month {}'.format(camera) + subfolder

            fits_file.writeto(full_folder_path+'.fits', overwrite=clobber)

def best_centered(fits_files, surround='neighbors', print_on=False):
    '''Given a list of FITS filenames, returns the file best centered PSF of the list.
    Best centered is defined as having the largest difference between the max pixel and its 4 surrounding pixels (top, bottom, left, right).'''
    hybrids_data = [fits.getdata(file) for file in fits_files]
    # list of the differences between the maximum pixel and its surrounding pixels
    center_surroundings_diff = []

    # creates the surrounding types
    if surround == 'basic':
        center_offsets = np.array([[-1,0], [+1,0], [0,-1], [0,+1]])
    elif surround == 'neighbors':
        center_offsets = np.array([[-1,0], [+1,0], [0,-1], [0,+1], [-1,-1], [-1,+1],[+1,-1],[+1,+1]])

    for hybrid_number, data in enumerate(hybrids_data):
        max_pixel_location = np.array(np.unravel_index(data.argmax(), data.shape))
        surrounding_pixels = np.tile(max_pixel_location,(len(center_offsets),1)) + center_offsets
        fancy_rows, fancy_cols = [p[0] for p in surrounding_pixels], [p[1] for p in surrounding_pixels]
        surrounding_pixel_values = data[fancy_rows, fancy_cols]

        surrounding_meanDN = np.mean(surrounding_pixel_values)
        center_surroundings_diff.append(np.max(data) - surrounding_meanDN)
        if print_on:
            print ('File: {}\nSurrounding:{}\navg_surround: {}\nmax_pixel: {}\nCenter Surround Diff: {}\n\n'.format(fits_files[hybrid_number], surrounding_pixel_values, surrounding_meanDN, np.max(data),center_surroundings_diff[hybrid_number]))
    return fits_files[np.argmax(center_surroundings_diff)]

def find_best_file(month, camera, region, surround='basic', print_on=False):
    '''Returns the best centered file using best_centered() and the given dataset, camera, and region.'''
    assert month == 6 or month == 18, 'Not a valid month for NavCam images'
    assert camera == 'NCM' or camera == 'NFT', 'Invalid camera'
    assert region in tagcams.region_colors.keys(), 'Invalid region name'

    filelist = glob.glob('hybrids data/L+{0}Month {1}/*/*/{2}.fits'.format(str(month), camera, region))
    if not filelist: # makes sure that if there is nothing in this filelist, that it doesn't get sent to the best_centered function
        return
    return best_centered(filelist, surround=surround, print_on=print_on)

def detect_centering_filechange(month, camera):
    for region in osiris18month.region_colors.keys():
        basic=find_best_file(month, camera, region, surround='basic', print_on=False)
        neighbors=find_best_file(month, camera, region, surround='neighbors', print_on=False)
        if basic != neighbors:
            print ('Change detected in {}'.format(region))
            print ('Basic: {}\nNeighbors:{}\n\n'.format(basic,neighbors))

def plot_from_region(month, camera, region, surround, print_on=False):
    '''Plots all PSFs from the given dataset, camera, and region.'''
    assert month == 6 or month == 18, 'Not a valid month for NavCam images'
    assert camera == 'NCM' or camera == 'NFT', 'Invalid camera'
    assert region in tagcams.region_colors.keys(), 'Invalid region name'

    filelist = glob.glob('hybrids data/L+{0}Month {1}/*/*/{2}.fits'.format(str(month), camera, region))
    if not filelist: # makes sure that if there is nothing in this filelist, that it doesn't get sent to the best_centered function
        return
    PSF_data = [fits.getdata(PSF_file) for PSF_file in filelist]
    fig, ax = plt.subplots(1,len(filelist), figsize=(12,3))

    for i, PSF in enumerate(PSF_data):
        ax[i].imshow(PSF, vmin=np.min(PSF_data), vmax=np.max(PSF_data))
        ax[i].grid(False), ax[i].set_title(filelist[i][filelist[i].find('gain'):], fontsize=8)
    fig.tight_layout()
    return best_centered(filelist, surround=surround, print_on=print_on)

def region_best_centered(month, camera, surround='basic'):
    '''Returns a dictionary mapping the regions to the filename of the best centered PSF.'''
    centered_PSFs_by_region = {}
    for region in tagcams.region_colors.keys():
        centered_PSFs_by_region[region] = find_best_file(month, camera, region, surround=surround)
    return centered_PSFs_by_region

def normalize_by_gain(month, camera, surround='basic'):
    '''Takes in a month and a camera value and computes the normalized PSFs by gain for each of these best centered PSFs.'''
    centered_PSFs_by_region = region_best_centered(month, camera, surround=surround)
    normalized_PSFs_gain = {}
    original_PSFs = {}

    for region in centered_PSFs_by_region.keys():
        if centered_PSFs_by_region[region] is not None:
            gain_word_loc = centered_PSFs_by_region[region].find('gain')
            if '8' in centered_PSFs_by_region[region][gain_word_loc+5:gain_word_loc+7]:
                gain = 1
            elif '10' in centered_PSFs_by_region[region][gain_word_loc+5:gain_word_loc+7]:
                gain = 1.25
            else:
                return 'There is a problem matching gains in the string'

            original_PSFs[region] = fits.getdata(centered_PSFs_by_region[region])
            normalized_PSFs_gain[region] = fits.getdata(centered_PSFs_by_region[region]) / gain

    return normalized_PSFs_gain, original_PSFs, centered_PSFs_by_region

def gaussian2d_fit(data, initial_parameters=True):
    '''Fits a 2D Gaussian using astropy's 2D Gaussian model and a Least Squares fitter (don't know which one is best)
    Returns a 10x10 array of a 2D Gaussian using the fitted parameters (amplitude, rotation, mean_x, mean_y, std_x, std_y)'''
    # picks which fitter to use—all are astropy fitters that use a combination of stats and scipy optimization methods
    fitter = fitting.LevMarLSQFitter()
    # creates meshgrid for all different combinations of x and y but retains the 10x10 hybrid shape
    x,y = np.mgrid[:data.shape[0], :data.shape[1]]

    # creates the moments function to get an initial guess at the fitting parameters
    def moments(data):
        """Returns (height, x, y, width_x, width_y)
        the gaussian parameters of a 2D distribution by calculating its
        moments """
        total = np.sum(data)
        X, Y = np.indices(data.shape)
        x = (X*data).sum()/total
        y = (Y*data).sum()/total
        col = data[:, int(y)]
        width_x = np.sqrt(abs((np.arange(col.size)-y)**2*col).sum()/col.sum())

        row = data[int(x), :]
        width_y = np.sqrt(abs((np.arange(row.size)-x)**2*row).sum()/row.sum())

        height = data.max()
        return height, x, y, width_x, width_y, 0.0

    if initial_parameters:
        height_guess, x_guess, y_guess, sigma_x_guess, sigma_y_guess = moments(data)[:-1]
        gaussian_model = models.Gaussian2D(amplitude=height_guess, x_mean=x_guess, y_mean=y_guess, x_stddev=sigma_x_guess, y_stddev=sigma_y_guess)
        first_fit = fitter(gaussian_model, x, y, z=data, maxiter=1000)
        initial_rotation = first_fit.theta.value
        # resets the initial theta value if it is out of the domain [0,2Pi]
        num_whole_rotations = initial_rotation // (2*np.pi)
        adjusted_rotation = initial_rotation - (2*np.pi) * num_whole_rotations

        first_fit_parameters = [first_fit.amplitude.value, first_fit.x_mean.value, first_fit.y_mean.value, first_fit.x_stddev.value, first_fit.y_stddev.value, adjusted_rotation]
        second_fit_model = models.Gaussian2D(*first_fit_parameters)
        final_fit = fitter(second_fit_model, x, y, z=data, maxiter=1000)
    elif initial_parameters == False:
        gaussian_model = models.Gaussian2D()
        final_fit = fitter(gaussian_model, x, y, z=data, maxiter=1000)

    return final_fit

def complicated_fit(original_data, size=[100,100]):
    '''Resamples the data to 100x100 (default) and computes the gaussfit from there'''
    upsampled = transform.resize(original_data, size, order=3, mode='reflect', anti_aliasing=True)
    upsampled_fit_array = gaussian2d_fit(upsampled)(*np.mgrid[:100,:100])
    downsampled = transform.resize(upsampled_fit_array, [10,10], order=3, mode='reflect', anti_aliasing=True)
    return gaussian2d_fit(downsampled)

def fit_regions(month, camera, surround='basic'):
    '''Spits out a bunch of dictionaries containing the parameters from a 2D Gaussian fit of the PSFs.'''
    normalized_PSFs_gain, centered_PSFs_by_region = normalize_by_gain(month, camera, surround=surround)[0::2]
    # initializes dictionaries that will store, for example, sigma_x = {'region': 37.2}
    distances = {}
    fitted_PSFs = {}
    amplitudes, sigma_x, sigma_y, rotations = {}, {}, {}, {}

    # get the defined center pixel of the dataset
    from ast import literal_eval as make_tuple # import to turn strings into tuples
    center_pixel = np.array(make_tuple(fits.getheader(centered_PSFs_by_region['center']).comments['Brightest Pixel of Target (Python) (5s)']))

    for region in normalized_PSFs_gain.keys():
        if normalized_PSFs_gain[region] is not None:
            # first get the filename and the distance from the center
            filename = centered_PSFs_by_region[region]
            target_pixel = np.array(make_tuple(fits.getheader(filename).comments['Brightest Pixel of Target (Python) (5s)']))

            # calculate its distance from the center
            #distance_from_center = np.linalg.norm([1436,1026]-target_pixel)
            distance_from_center = np.linalg.norm(center_pixel-target_pixel)
            distances[region] = distance_from_center

            # fit each PSF to the Gaussian 2D and extract parameters into the dictionaries
            x,y = np.mgrid[:10, :10]
            '''Why can't I catch this warning about fitting the Gaussian?!'''
            fitted_model = gaussian2d_fit(normalized_PSFs_gain[region])
            fitted_PSFs[region] = fitted_model(x,y)
            amplitudes[region] = fitted_model.amplitude
            sigma_x[region], sigma_y[region] = fitted_model.y_stddev, fitted_model.x_stddev # switched order due to Python's (y,x) convention
            rotations[region] = fitted_model.theta.value

    return distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs

def fit_PSFs_to_excel(month, camera, surround='basic', excel_name=None):
    '''Puts the fit_regions information in an Excel file and returns a DataFrame'''
    distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs = fit_regions(month, camera, surround)
    df =  pd.DataFrame(columns=['Distance', 'Peak Height', 'x width', 'y width', 'Filename', 'Rotation'])
    coordinates = []
    for region in distances.keys():
        df.loc[region] = [distances[region], amplitudes[region][0], sigma_x[region][0], sigma_y[region][0], centered_PSFs_by_region[region], rotations[region]]
        coordinate_string = fits.getheader(centered_PSFs_by_region[region]).comments['Brightest Pixel of Target (Python) (5s)']

        from ast import literal_eval as make_tuple # import to turn strings into tuples
        pixel_coordinate = make_tuple(coordinate_string)
        coordinates.append(pixel_coordinate)
    df['Pixel Coordinates (Python)'] = coordinates
    if excel_name is not None:
        with pd.ExcelWriter(excel_name) as writer:
            df.to_excel(writer)
    return df

def overwrite_excel_stats(new_excel):
    df_6month_NCM = fit_PSFs_to_excel(6, 'NCM')
    df_6month_NFT = fit_PSFs_to_excel(6, 'NFT')
    df_18month_NCM = fit_PSFs_to_excel(18, 'NCM')
    df_18month_NFT = fit_PSFs_to_excel(18, 'NFT')
    df_6months = pd.concat([df_6month_NCM, df_6month_NFT])
    df_18months = pd.concat([df_18month_NCM, df_18month_NFT])

    from openpyxl import load_workbook
    existing_workbook = load_workbook('statistics for fits no upsampling removed center outlier.xlsx')

    with pd.ExcelWriter(new_excel, engine='openpyxl') as writer:
        writer.book = existing_workbook
        df_6months.to_excel(writer, sheet_name='L+6Months Gaussian Parameters')
        df_18months.to_excel(writer, sheet_name = 'L+18Months Gaussian Parameters')

def corner_fitted_and_original(month, camera, corner, surround='basic'):
    '''Simple function to plot all the original and fitted PSFs side by side for each region for comparison purposes.'''
    distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs = fit_regions(month, camera, surround=surround)

    f, [ax1,ax2,ax3] = plt.subplots(1,3, figsize=(10,4))
    ax1.grid(False), ax2.grid(False), ax3.grid(False), ax1.set_title('Original PSF'), ax2.set_title('Fitted Gaussian'), ax3.set_title('Residuals')
    ax1.imshow(normalized_PSFs_gain[corner])
    ax2.imshow(fitted_PSFs[corner], vmin=np.min(normalized_PSFs_gain[corner]), vmax=np.max(normalized_PSFs_gain[corner]))
    ax3.imshow(normalized_PSFs_gain[corner]-fitted_PSFs[corner], vmin=np.min(normalized_PSFs_gain[corner]), vmax=np.max(normalized_PSFs_gain[corner]))

    f.suptitle('{}\n{}'.format(corner, centered_PSFs_by_region[corner]))
    f.tight_layout(), f.subplots_adjust(top=.85)

def PSFs_original_and_fitted(month, camera, surround='basic'):
    '''Simple function to plot all the original and fitted PSFs side by side for each region for comparison purposes.'''
    distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs = fit_regions(month, camera, surround=surround)
    figures = []
    corners = []
    for corner in normalized_PSFs_gain.keys():
        f, [ax1,ax2,ax3] = plt.subplots(1,3, figsize=(10,4))
        ax1.grid(False), ax2.grid(False), ax3.grid(False), ax1.set_title('Original PSF'), ax2.set_title('Fitted Gaussian'), ax3.set_title('Residuals')
        ax1.imshow(normalized_PSFs_gain[corner])
        ax2.imshow(fitted_PSFs[corner], vmin=np.min(normalized_PSFs_gain[corner]), vmax=np.max(normalized_PSFs_gain[corner]))
        ax3.imshow(normalized_PSFs_gain[corner]-fitted_PSFs[corner], vmin=np.min(normalized_PSFs_gain[corner]), vmax=np.max(normalized_PSFs_gain[corner]))

        f.suptitle('{}\n{}'.format(corner, centered_PSFs_by_region[corner]))
        f.tight_layout(), f.subplots_adjust(top=.85)
        figures.append(f)
        corners.append(corner)
    return figures, corners

def fitted_in_powerpoint(month, camera):
    figures, corners = PSFs_original_and_fitted(month, camera)
    for fig, corner in zip(figures, corners):
        fig.savefig('original and fitted hybrid PSFs/L+{}Months/{}/{}'.format(month, camera, corner), dpi=500)

def remove_outliers_mask(dataset, m=2):
    '''Gets rid of outliers that are 1.5 standard deviations from the median of the data.
    Input has to be an array, also returns an array of the MASK that will remove outliers, not the array itself.'''
    assert type(dataset) == np.ndarray, 'Input data is not in array format'
    return np.abs(dataset - np.median(dataset)) < m * np.std(dataset)
'''from scipy import misc
x,y=np.mgrid[:10,:10]
simulated_data = misc.imread('PSF_OnAxis_sys_10x10subwindow.bmp',flatten=1)
plt.imshow(simulated_data)
plt.imshow(gaussian2d_fit(simulated_data)(x,y))'''
#gaussian2d_fit(simulated_data)
def linear_regression(x,y):
    '''Produces line of best fit according to x and y values'''
    lin_reg = LinearRegression()
    return lin_reg.fit(x.reshape(-1,1),y.reshape(-1,1))

def stat_vs_distance(month, outliers=None, surround='basic', annotate=True):
    '''Plots the peak height, x-width, and y-width as a function of radial distance. Pass in the month
    and the outliers, which is a list of lists corresponding to NCM and NFT outliers, respectively.'''
    fig, ax_stats = plt.subplots(3,1,figsize=(11,9), sharex=True)
    PSF_fits_bothcameras = []
    distances_bothcameras = []
    normed_PSFs = []

    # create a pandas dataframe containing the statistics from the regression fits
    regression_stats = pd.DataFrame(columns=['Slope', 'Intercept', 'Coefficient of determination (R^2)'])

    for camera_index, camera in enumerate(['NCM', 'NFT']):
        # load previous data
        distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs = fit_regions(month, camera, surround=surround)
        PSF_fits_bothcameras.append(fitted_PSFs), distances_bothcameras.append(distances), normed_PSFs.append(normalized_PSFs_gain)
        gaussian_parameters = [amplitudes, sigma_x, sigma_y]
        stat_names = ['Peak Height', r'$\mathrm{\sigma_x}$', r'$\mathrm{\sigma_y}$']
        stat_names_nice_string = ['Peak Height', 'x width', 'y width']

        for ax_number, p in enumerate(gaussian_parameters): # plot stats on separate plots
            active_axis = ax_stats[ax_number]
            '''outliers_mask = remove_outliers_mask(np.array([i.value for i in p.values()]))
            x_vals = np.array([d for d in distances.values()])
            y_vals = np.array([i.value for i in p.values()])
            x_vals, y_vals= x_vals.flatten()[outliers_mask], y_vals.flatten()[outliers_mask]'''
            x_vals, y_vals = np.array([]), np.array([])

            for region, dist in distances.items(): # ensures that the x_vals and y_vals are populated in the correct order
                if outliers is not None:
                    if region not in outliers[camera_index]:
                        x_vals = np.append(x_vals, dist)
                        y_vals = np.append(y_vals, p[region])
                        if annotate:
                            active_axis.annotate(region, (dist, p[region].value))
                elif outliers is None:
                    x_vals = np.append(x_vals, dist)
                    y_vals = np.append(y_vals, p[region])
                    if annotate:
                        active_axis.annotate(region, (dist, p[region].value))

            active_axis.scatter(x_vals, y_vals, label=camera) # scatter plot of the x and y points

            # computes the regression line for the data and plots it
            reg_line = linear_regression(x_vals, y_vals)
            slope, intercept = reg_line.coef_, reg_line.intercept_
            assert len(slope) == 1 and len(intercept) == 1
            slope, intercept = reg_line.coef_[0][0], reg_line.intercept_[0]
            y_predicted = reg_line.predict(x_vals.reshape(-1,1)).flatten()
            active_axis.plot(x_vals, y_predicted)
            active_axis.set(ylabel=stat_names[ax_number])

            from sklearn.metrics import r2_score as r2
            correlation_coefficient = r2(y_vals, y_predicted)
            print ('{} {} R-squared: {}'.format(camera, stat_names_nice_string[ax_number], correlation_coefficient))

            # compile all the stats for one iteration into a dataframe to be appended to regression_stats
            row_stat = pd.DataFrame({'Slope': slope, 'Intercept': intercept, 'Coefficient of determination (R^2)': correlation_coefficient},
                                    index=['{} {}'.format(camera, stat_names_nice_string[ax_number])],
                                    columns=['Slope', 'Intercept', 'Coefficient of determination (R^2)'])

            regression_stats = regression_stats.append(row_stat)

    ax_stats[0].set_title('L+{}Months'.format(month)), ax_stats[0].legend()
    ax_stats[-1].set_xlabel('Distance from Center')

    fig.tight_layout()
    fig.subplots_adjust(hspace=0.05)

    return fig, regression_stats

def excel_stats(excel_name, six_outliers=None, eighteen_outliers=None):
    '''Creates an Excel file based on the statistics of the regression line fits.'''
    stats_6month = stat_vs_distance(6, outliers=six_outliers, surround='basic')[1]
    stats_18month = stat_vs_distance(18, outliers=eighteen_outliers, surround='basic')[1]
    combined_stats_basic = stats_6month.append(stats_18month)

    stats_6month_neighbors = stat_vs_distance(6, outliers=six_outliers, surround='neighbors')[1]
    stats_18month_neighbors = stat_vs_distance(18, outliers=eighteen_outliers, surround='neighbors')[1]
    combined_neighbors = stats_6month_neighbors.append(stats_18month_neighbors)
    with pd.ExcelWriter(excel_name) as writer:
        combined_stats_basic.to_excel(writer, sheet_name='Basic Centering')
        combined_neighbors.to_excel(writer, sheet_name='Neighbors')

def error_bars(month, camera):
    '''For a given month and camera value, calculate the error for each of the total PSF values.'''
    error_by_region, PSF_totals_by_region = {}, {}
    for region in osiris18month.region_colors.keys():
        filelist = glob.glob('hybrids data/L+{0}Month {1}/*/*/{2}.fits'.format(str(month), camera, region))
        if not filelist: # makes sure that if there is nothing in this filelist, that it doesn't get sent to the best_centered function
            continue

        gain_word_loc = [filename.find('gain') for filename in filelist]
        gain_array = np.array([1 if '8' in filename[gain_word_loc[i]+5:gain_word_loc[i]+7] else 1.25 for i, filename in enumerate(filelist)])
        assert len(gain_array) == len(filelist), 'Number of gains not matching filelist length'

        total_PSF_data = np.array([np.sum(fits.getdata(PSF_file)/gain_array[i]) for i, PSF_file in enumerate(filelist)])
        bootstrapped_SD = bootstrap(total_PSF_data) # bootstraps the total PSF data and returns a bootstrapped sample
        three_sigma = 3*np.mean(bootstrapped_SD) # gets the value of 3 standard deviations for error bar
        error_by_region[region] = three_sigma
        PSF_totals_by_region[region] = total_PSF_data

    return error_by_region

def bootstrap(initial_dataset, bootstrap_size=3, num_resamples=10000, stat=np.std):
    '''Bootstraps a statistic, default stat is the standard deviation'''
    bootstrapped = [stat(np.random.choice(initial_dataset,size=bootstrap_size, replace=True)) for i in np.arange(num_resamples)]
    return bootstrapped

def PSF_total_energy(month, fitted=False, mask=None, plot_errors=True, regression_line=False):
    PSF_fits_bothcameras, distances_bothcameras, normed_PSFs = [], [], []

    for camera in ['NCM', 'NFT']:
        # load previous data
        distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs = fit_regions(month, camera)
        PSF_fits_bothcameras.append(fitted_PSFs), distances_bothcameras.append(distances), normed_PSFs.append(normalized_PSFs_gain)

    fig, ax = plt.subplots(1,1, figsize=(10,7))
    if fitted == False:
        PSFs_to_use = normed_PSFs
        title = 'Original'
    elif fitted == True:
        PSFs_to_use = PSF_fits_bothcameras
        title = 'Fitted Gaussian'
    else:
        return 'Not a valid entry for parameter "fitted"'

    for i, camera in enumerate(['NCM', 'NFT']):
        radial_distances, energies, errors = [], [], []
        measurement_errors = error_bars(month, camera) # gets dictionary of error for each energy
        #return distances_bothcameras[i], {region:np.sum(PSFs_to_use[i][region]) for region in PSFs_to_use[i].keys()}, measurement_errors
        assert distances_bothcameras[i].keys() == measurement_errors.keys() and distances_bothcameras[i].keys() == PSFs_to_use[i].keys()

        for region in distances_bothcameras[i].keys():
            if PSFs_to_use[i][region] is not None:
                radial_distance = distances_bothcameras[i][region]
                total_energy = np.sum(PSFs_to_use[i][region])
                radial_distances.append(radial_distance), energies.append(total_energy), errors.append(measurement_errors[region])
                ax.annotate(region, (radial_distance, total_energy))
        if mask is not None:
            mask = remove_outliers_mask(energies)

        radial_distances = np.array(radial_distances)[mask].flatten()
        energies = np.array(energies)[mask].flatten()

        if plot_errors == True:
            ax.errorbar(radial_distances, energies, yerr=errors, fmt='o', alpha=0.6)
            ax.scatter(radial_distances, energies, label=camera)
        else:
            ax.scatter(radial_distances, energies, label=camera)

        if regression_line:
            # get the linear regression prediction
            reg_line = linear_regression(radial_distances, energies)
            y_predicted = reg_line.predict(radial_distances.reshape(-1,1)).flatten()
            correlation_coefficient = reg_line.score(y_predicted.reshape(-1,1), radial_distances.reshape(-1,1))
            print ('{} r-squared: {}'.format(camera, correlation_coefficient))
            ax.plot(radial_distances.flatten(), y_predicted)

    ax.legend()
    ax.set(title='L+{}Months {} PSF Energies'.format(month, title), xlabel='Distance from Center', ylabel='Total DN/s')
    fig.tight_layout()

    return fig, distances_bothcameras, PSFs_to_use

def stats_polyfit(month, degree=1):
    fig, ax_stats = plt.subplots(3,1,figsize=(11,9), sharex=True)
    PSF_fits_bothcameras = []
    distances_bothcameras = []
    normed_PSFs = []
    for camera in ['NCM', 'NFT']:
        # load previous data
        distances, amplitudes, sigma_x, sigma_y, rotations, [normalized_PSFs_gain, centered_PSFs_by_region], fitted_PSFs = fit_regions(month, camera)
        PSF_fits_bothcameras.append(fitted_PSFs), distances_bothcameras.append(distances), normed_PSFs.append(normalized_PSFs_gain)
        gaussian_parameters = [amplitudes, sigma_x, sigma_y]
        stat_names = ['Peak Height', r'$\mathrm{\sigma_x}$', r'$\mathrm{\sigma_y}$']
        for ax_number, p in enumerate(gaussian_parameters): # plot stats on separate plots
            active_axis = ax_stats[ax_number]
            outliers_mask = remove_outliers_mask(np.array([i.value for i in p.values()]))
            x_vals = np.array([d for d in distances.values()])
            y_vals = np.array([i.value for i in p.values()])
            x_vals, y_vals= x_vals[outliers_mask], y_vals[outliers_mask]

            active_axis.scatter(x_vals, y_vals, label=camera) # scatter plot of the x and y points

            # computes the regression line for the data and plots it
            fitted_line = np.polyfit(x_vals, y_vals, degree)
            y_predicted = np.poly1d(fitted_line)(x_vals)
            x_sorted, y_sorted_predicted = zip(*sorted(zip(x_vals, y_predicted)))
            active_axis.plot(x_sorted, y_sorted_predicted)
            active_axis.set(ylabel=stat_names[ax_number]), active_axis.legend()

            '''regions_for_annotation = np.array([k for k in p.keys()])[outliers_mask]
            for i, r in enumerate(regions_for_annotation):
                active_axis.annotate(r, (x_vals[outliers_mask][i], y_vals[outliers_mask][i]))'''

    ax_stats[0].set_title('L+{}Months'.format(month))
    ax_stats[-1].set_xlabel('Distance from Center')

    fig.tight_layout()
    fig.subplots_adjust(hspace=0.05)

    return fig, [x_vals, y_vals, y_predicted]

def residuals_calculation(true_values, predicted):
    '''Calculates residuals.'''
    if type(predicted) != np.ndarray:
        predicted = np.array(predicted)
    if type(true_values) != np.ndarray:
        true_values = np.array(true_values)
    return true_values - predicted

def create_pixelcenter_dictionary(month, camera):
    '''Given a month and a camera, returns a dictionary of the best centered target locations and their respective regions'''
    normalized_PSFs_gain, centered_PSFs_by_region = normalize_by_gain(month, camera)[0::2]

    # get the defined center pixel of the dataset from centered_PSFs_by_region
    from ast import literal_eval as make_tuple # import to turn strings into tuples
    center_pixel = np.array(make_tuple(fits.getheader(centered_PSFs_by_region['center']).comments['Brightest Pixel of Target (Python) (5s)']))
    regions_and_targetlocations = {region:np.array(make_tuple(fits.getheader(centered_PSFs_by_region[region]).comments['Brightest Pixel of Target (Python) (5s)']))\
    for region in centered_PSFs_by_region.keys() if centered_PSFs_by_region[region] is not None}
    return regions_and_targetlocations

def rectangles_bestcentered(PSF_centers_dictionary, x_location=[70,115], y_location=[163,38], origin_location=[15,45], box_dimensions=[130,100], vmin_vmax=[0,30],output_name=None):
    '''Creates a plot of with a bunch of rectangular regions that represent PSF regions.
    Also indicates the coordinate system that will be used for the presentation.
    For standard rectangles, box_dimensions=[130,100]'''

    fig, ax = plt.subplots(1,1,figsize=(10,8))
    ax.grid(False)

    # create a rectangle 2752 x 2004 image rectangle representing a sample image
    sample_image_file = osiris18month.five_second_images[31]
    ax.imshow(fits.getdata(sample_image_file),vmin=vmin_vmax[0],vmax=vmin_vmax[1],cmap='gray')

    for region in PSF_centers_dictionary.keys():
        brightest_column, brightest_row = PSF_centers_dictionary[region] # flipped because the Excel file is written in (x,y) format
        print ('Assigning region to {}'.format(region))
        assigned_region = osiris18month.assign_region([brightest_row, brightest_column])
        rect_color = osiris18month.region_colors[assigned_region]
        rectangle = patches.Rectangle((brightest_column-(box_dimensions[0]/2),brightest_row-(box_dimensions[1]/2)), box_dimensions[0],box_dimensions[1], edgecolor=rect_color,linewidth=0.7,facecolor='none')
        _= ax.add_patch(rectangle)
    ax.arrow(134,50,100,0, color='red', head_width=15)
    ax.arrow(134,50,0,100, color='red', head_width=15)
    ax.set(xticks=[],yticks=[])
    ax.annotate('+X', xy=x_location, xycoords='data', color='red', fontname='Myriad Pro')
    ax.annotate('+Y', xy=y_location, xycoords='data', color='red', fontname='Myriad Pro')
    ax.annotate('(0,0)', xy=origin_location, xycoords='data', color='white', fontname='Myriad Pro')
    fig.tight_layout()

    if output_name is not None:
        with PdfPages(output_name) as pdf:
            pdf.savefig(fig, dpi=1000)
    return fig, ax
